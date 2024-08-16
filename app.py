from flask import Flask, render_template, request, jsonify, redirect, url_for, session
import face_recognition
import cv2
import numpy as np
from base64 import b64decode
from io import BytesIO
import openpyxl
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Load names and image paths from Excel
data_dir = 'faces'
data_file = os.path.join(data_dir, 'faces.xlsx')

# Ensure the images directory exists
image_folder = 'images'
if not os.path.exists(image_folder):
    os.makedirs(image_folder)

def load_faces_df():
    try:
        workbook = openpyxl.load_workbook(data_file)
        sheet = workbook.active
        faces_data = []
        for row in sheet.iter_rows(values_only=True):
            faces_data.append(row)
        return faces_data[1:]  # Skip header row
    except FileNotFoundError:
        return []

def save_faces_df(faces_data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Name', 'Image_Path'])  # Header row
    for row in faces_data:
        sheet.append(row)
    workbook.save(data_file)

@app.route('/process_image', methods=['POST'])
def process_image():
    if not session.get('verified'):
        return jsonify({"message": "Unauthorized access"}), 401

    try:
        data = request.json['image']
        # Decode the image data
        header, encoded = data.split(",", 1)
        binary_data = b64decode(encoded)
        image_np = np.frombuffer(binary_data, dtype=np.uint8)
        image = cv2.imdecode(image_np, cv2.IMREAD_COLOR)

        # Convert image to RGB (if necessary)
        rgb_frame = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        # Process the image
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        faces_data = load_faces_df()
        known_face_encodings = []
        known_face_names = []

        for name, img_path in faces_data:
            known_image = face_recognition.load_image_file(os.path.join(image_folder, img_path))
            known_face_encoding = face_recognition.face_encodings(known_image)[0]
            known_face_encodings.append(known_face_encoding)
            known_face_names.append(name)

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)

            if matches[best_match_index]:
                name = known_face_names[best_match_index]
                return jsonify({"match": name, "confidence": 1 - face_distances[best_match_index], "date": pd.Timestamp.now().strftime('%Y-%m-%d')})

        return jsonify({"message": "No matching face found"})

    except Exception as e:
        return jsonify({"message": f"Internal server error: {str(e)}"}), 500

@app.route('/take_photo', methods=['GET', 'POST'])
def take_photo():
    if request.method == 'POST':
        try:
            data = request.json['image']
            # Decode the image data
            header, encoded = data.split(",", 1)
            binary_data = b64decode(encoded)
            image_np = np.frombuffer(binary_data, dtype=np.uint8)
            image = cv2.imdecode(image_np, cv2.IMREAD_COLOR)

            # Resize image to make it square (if needed)
            h, w = image.shape[:2]
            desired_size = max(h, w)
            delta_w = desired_size - w
            delta_h = desired_size - h
            top, bottom = delta_h // 2, delta_h - (delta_h // 2)
            left, right = delta_w // 2, delta_w - (delta_w // 2)
            color = [0, 0, 0]
            image = cv2.copyMakeBorder(image, top, bottom, left, right, cv2.BORDER_CONSTANT, value=color)

            # Save the image with a unique name
            photo_count = len(os.listdir(image_folder))
            new_image_name = f"{photo_count + 1:04d}.jpg"  # Adding 1 to ensure a unique name
            cv2.imwrite(os.path.join(image_folder, new_image_name), image)

            # Load and update faces.xlsx
            faces_data = load_faces_df()
            faces_data.append([photo_count + 1, new_image_name])
            save_faces_df(faces_data)

            # Set session as verified
            session['verified'] = True

            return jsonify({"message": "Photo taken and saved", "image_name": new_image_name})
        except Exception as e:
            return jsonify({"message": f"Internal server error: {str(e)}"}), 500

    return render_template('take_photo.html')

@app.route('/direct_face_recognition')
def direct_face_recognition():
    # Directly redirect to face recognition without QR scanning
    session['verified'] = True  # Assuming direct recognition should skip any verification
    return redirect(url_for('face_recognition'))

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        qr_code = request.get_json().get('code')
        # Here should be the code to validate QR and respond with relevant redirection
        # Assuming verification logic here...

        return jsonify({"redirect": url_for('take_photo')})

    return render_template('index.html')

@app.route('/qr_scan')
def qr_scan():
    # Render the QR scanning interface
    return render_template('qr_scan.html')

@app.route('/face_recognition')
def face_recognition():
    if not session.get('verified'):
        return redirect(url_for('index'))
    return render_template('face_recognition.html')

if __name__ == '__main__':
    app.run(debug=True)
